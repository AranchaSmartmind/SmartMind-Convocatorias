"""
Módulo de procesamiento de datos
Extrae información de PDFs y Excel
"""

import re
import pdfplumber
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
from collections import defaultdict
import shutil
import tempfile

def obtener_mes_anterior():
    """
    Obtiene el nombre del mes anterior al actual
    
    Returns:
        str: Nombre del mes en mayúsculas
    """
    mes_anterior = datetime.now() - relativedelta(months=1)
    meses = {
        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
        5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
        9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
    }
    return meses[mes_anterior.month]

def extraer_datos_curso_pdf(pdf_path):
    """
    Extrae datos básicos del curso del PDF de becas
    
    Args:
        pdf_path: Ruta al PDF de becas
    
    Returns:
        tuple: (numero_curso, especialidad)
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()
            
            # Número de curso
            numero_curso = "2024/1339"
            match = re.search(r'Nº de Curso\s+(\d{4}/\d{4})', text)
            if match:
                numero_curso = match.group(1)
            
            # Especialidad
            especialidad = "OPERACIONES AUXILIARES DE SERVICIOS ADMINISTRATIVOS Y GENERALES"
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if 'OPERACIONES AUXILIARES' in line:
                    especialidad = line.strip()
                    break
            
            return numero_curso, especialidad
    except Exception as e:
        print(f"Error extrayendo datos del curso: {e}")
        return "2024/1339", "OPERACIONES AUXILIARES DE SERVICIOS ADMINISTRATIVOS Y GENERALES"

def extraer_alumnos_excel(excel_path):
    """
    Extrae listado de alumnos del Excel
    
    Args:
        excel_path: Ruta al archivo Excel
    
    Returns:
        list: Lista de diccionarios con datos de alumnos
    """
    import shutil
    import tempfile
    
    # Crear una copia temporal del archivo
    temp_file = None
    wb = None
    
    try:
        # Crear archivo temporal con un nombre único
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
            temp_file = tmp.name
            # Copiar contenido
            with open(excel_path, 'rb') as source:
                shutil.copyfileobj(source, tmp)
        
        # Ahora trabajar con la copia temporal
        wb = load_workbook(temp_file, read_only=True, data_only=True)
        ws = wb.active
        
        alumnos = []
        header_row = None
        
        # Buscar fila de encabezado
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
            for cell in row:
                if cell.value and 'APELLIDOS, NOMBRE' in str(cell.value):
                    header_row = i
                    break
            if header_row:
                break
        
        if not header_row:
            return []
        
        # Leer datos
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row[1] and isinstance(row[1], str) and ',' in row[1]:
                alumnos.append({
                    'nombre_completo': row[1].strip(),
                    'dni': str(row[2]) if row[2] else ''
                })
        
        return alumnos
        
    except Exception as e:
        print(f"Error leyendo Excel: {e}")
        return []
        
    finally:
        # Cerrar workbook
        if wb:
            try:
                wb.close()
            except:
                pass
        
        # Eliminar archivo temporal
        if temp_file:
            try:
                import os
                import time
                time.sleep(0.1)  # Pequeña pausa para asegurar que se liberó
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except:
                pass

def extraer_becas_ayudas_simple(pdf_path):
    """
    Extrae becas y ayudas del PDF
    
    Args:
        pdf_path: Ruta al PDF de becas
    
    Returns:
        dict: Diccionario {nombre: [lista de ayudas]}
    """
    ayudas_por_alumno = {}
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            tables = pdf.pages[0].extract_tables()
            
            if tables and tables[0]:
                table = tables[0]
                
                # Buscar encabezados
                header_row = None
                for i, row in enumerate(table):
                    if any('Apellidos' in str(cell) or 'Nombre' in str(cell) for cell in row if cell):
                        header_row = i
                        break
                
                if not header_row:
                    return ayudas_por_alumno
                
                headers = table[header_row]
                
                # Mapear columnas
                col_indices = {}
                for i, header in enumerate(headers):
                    if not header:
                        continue
                    header_lower = str(header).lower()
                    if 'apellidos' in header_lower or 'nombre' in header_lower:
                        col_indices['nombre'] = i
                    elif 'transporte' in header_lower:
                        col_indices['transporte'] = i
                    elif 'conciliación' in header_lower:
                        col_indices['conciliacion'] = i
                    elif 'beca' in header_lower:
                        col_indices['beca'] = i
                    elif 'manutención' in header_lower:
                        col_indices['manutencion'] = i
                    elif 'discapacidad' in header_lower:
                        col_indices['discapacidad'] = i
                
                # Procesar filas
                for row in table[header_row + 1:]:
                    if not row or not row[col_indices.get('nombre', 0)]:
                        continue
                    
                    nombre = str(row[col_indices['nombre']]).strip()
                    if not nombre or len(nombre) < 3:
                        continue
                    
                    ayudas = []
                    
                    for tipo, idx in col_indices.items():
                        if tipo == 'nombre':
                            continue
                        
                        if idx >= len(row):
                            continue
                        
                        valor = str(row[idx]) if row[idx] else ''
                        
                        if 'X' in valor.upper() or 'x' in valor:
                            if tipo == 'transporte':
                                ayudas.append('Transporte')
                            elif tipo == 'conciliacion':
                                ayudas.append('conciliación')
                            elif tipo == 'beca':
                                ayudas.append('Beca')
                            elif tipo == 'manutencion':
                                ayudas.append('Manutención')
                            elif tipo == 'discapacidad':
                                ayudas.append('Discapacidad')
                    
                    if ayudas:
                        ayudas_por_alumno[nombre] = ayudas
    
    except Exception as e:
        print(f"Error extrayendo becas: {e}")
    
    return ayudas_por_alumno

def extraer_justificantes(pdf_path):
    """
    Extrae justificantes del PDF
    
    Args:
        pdf_path: Ruta al PDF de justificantes
    
    Returns:
        dict: Diccionario {nombre: cantidad}
    """
    justificantes = defaultdict(int)
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                
                if 'JUSTIFICANTE' in text or 'prestación servicios' in text:
                    patrones = [
                        r'(?:trabajadora|paciente|alumno|Doña|Don)\s+(?:Doña|Don)?\s*([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)+)',
                        r'([A-Z][A-ZÁÉÍÓÚÑ]+(?:\s+[A-Z][A-ZÁÉÍÓÚÑ]+)+)[,\s]+(?:con|NIE|DNI)'
                    ]
                    
                    for patron in patrones:
                        match = re.search(patron, text, re.IGNORECASE)
                        if match:
                            nombre = match.group(1).strip().upper()
                            justificantes[nombre] += 1
                            break
    except Exception as e:
        print(f"Error extrayendo justificantes: {e}")
    
    return justificantes