"""
Sección de Formación Empresa Fin
"""
import streamlit as st
import io
import pandas as pd
from datetime import datetime
from utils import (
    procesar_documento,
    leer_datos_ctrl,
    leer_datos_excel,
    extraer_datos_multiples_documentos
)


def extraer_datos_certificado_asistencia(texto):
    """Extrae nombres y DNIs de la Hoja de Firmas"""
    import re
    
    datos = {
        "alumnos": [],
        "fecha_inicio": "",
        "curso": ""
    }

    match_curso = re.search(r'Especialidad[:\s]+([^\n\(]+)', texto, re.IGNORECASE)
    if match_curso:
        datos["curso"] = match_curso.group(1).strip()

    fechas = re.findall(r'(\d{2}/\d{2}/\d{4})', texto)
    if fechas:
        datos["fecha_inicio"] = fechas[0]

    patron = r'(\d{8}[A-Z])\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+?)(?=\s*\d{8}[A-Z]|\n\n|LUNES|MARTES|\Z)'
    matches = re.findall(patron, texto, re.DOTALL)

    palabras_excluir = ['INTERPROS', 'GENERATION', 'OPERACIONES', 'AUXILIARES',
                        'SERVICIOS', 'ADMINISTRATIVOS', 'PRINCIPADO', 'ASTURIAS',
                        'PUBLICO', 'EMPLEO', 'ENTRADA', 'SALIDA', 'MIERCOLES']

    for dni, nombre in matches:
        nombre_limpio = ' '.join(nombre.strip().split())
        es_valido = True
        for palabra in palabras_excluir:
            if palabra in nombre_limpio.upper():
                es_valido = False
                break
        if es_valido and len(nombre_limpio) > 10:
            datos["alumnos"].append((nombre_limpio, dni))

    return datos


def extraer_evaluacion_profesores(texto):
    """Extrae nota final y calificación del documento Evaluación de Profesores"""
    import re
    
    datos = {
        "alumnos": {}
    }

    lineas = texto.split('\n')

    nombre_actual = None
    for i, linea in enumerate(lineas):
        linea = linea.strip()

        if ',' in linea and len(linea) > 10:
            partes = linea.split(',')
            if len(partes) == 2 and all(p.strip().replace(' ', '').isalpha() for p in partes):
                nombre_actual = linea.upper().strip()
                if nombre_actual not in datos["alumnos"]:
                    datos["alumnos"][nombre_actual] = {}

        if nombre_actual:
            linea_upper = linea.upper()

            if 'MF' in linea_upper and '_' in linea_upper:
                match_modulo = re.search(r'(MF\d{4}_\d)', linea_upper)
                if match_modulo:
                    modulo = match_modulo.group(1)

                    nota = None
                    calificacion = None

                    for j in range(i, min(len(lineas), i+6)):
                        linea_busqueda = lineas[j].strip()

                        match_nota = re.search(r'\b(\d{1,2}(?:[.,]\d{1,2})?)\b', linea_busqueda)
                        if match_nota and not nota:
                            nota_str = match_nota.group(1).replace(',', '.')
                            try:
                                nota_float = float(nota_str)
                                if 0 <= nota_float <= 10:
                                    nota = nota_str
                            except:
                                pass

                        linea_upper_busqueda = linea_busqueda.upper()
                        calificaciones = ['SUPERADO', 'NO SUPERADO', 'CONVALIDADO', 'CONVALIDA', 'APTO', 'NO APTO', 'EXENTO']
                        for calif in calificaciones:
                            if calif in linea_upper_busqueda and not calificacion:
                                calificacion = calif
                                break

                    if nota and calificacion:
                        valor_final = f"{nota} {calificacion}"
                        datos["alumnos"][nombre_actual][modulo] = valor_final
                    elif nota:
                        datos["alumnos"][nombre_actual][modulo] = nota
                    elif calificacion:
                        datos["alumnos"][nombre_actual][modulo] = calificacion

    return datos


def extraer_evaluacion_excel(file, verbose=False):
    """Lee el Excel de Evaluación buscando columnas NOTA FINAL y Superado dentro de cada módulo"""
    import re
    
    datos = {
        "alumnos": {}
    }

    try:
        file.seek(0)
        df_raw = pd.read_excel(file, sheet_name=0, header=None)

        fila_modulos = None
        for idx, row in df_raw.iterrows():
            row_text = ' '.join([str(x) for x in row.values if pd.notna(x)])
            if 'MF0969' in row_text or 'MF0970' in row_text:
                fila_modulos = idx
                break

        if not fila_modulos:
            return datos

        col_nombres = None
        for col_idx in range(min(10, df_raw.shape[1])):
            for fila_idx in range(fila_modulos + 1, min(fila_modulos + 15, df_raw.shape[0])):
                valor = df_raw.iloc[fila_idx, col_idx]
                if pd.notna(valor) and ',' in str(valor) and len(str(valor)) > 10:
                    col_nombres = col_idx
                    break
            if col_nombres is not None:
                break

        if col_nombres is None:
            return datos

        fila_mod = df_raw.iloc[fila_modulos]
        modulos_info = []

        for col_idx, valor in enumerate(fila_mod):
            if pd.notna(valor) and 'MF' in str(valor).upper():
                match = re.search(r'(MF\d{4}_\d)', str(valor).upper())
                if match:
                    modulo = match.group(1)
                    modulos_info.append({"modulo": modulo, "col_inicio": col_idx})

        for i, info in enumerate(modulos_info):
            modulo = info["modulo"]
            col_inicio = info["col_inicio"]
            col_fin = modulos_info[i + 1]["col_inicio"] if i + 1 < len(modulos_info) else df_raw.shape[1]

            nota_col = None
            calif_col = None

            for col_idx in range(col_inicio, col_fin):
                for fila_enc in range(fila_modulos, min(fila_modulos + 5, df_raw.shape[0])):
                    celda = str(df_raw.iloc[fila_enc, col_idx])

                    if 'NOTA' in celda.upper() and 'FINAL' in celda.upper():
                        nota_col = col_idx
                    if 'SUPERADO' in celda.upper():
                        calif_col = col_idx

            if nota_col and calif_col:
                info["nota_col"] = nota_col
                info["calif_col"] = calif_col

        fila_inicio = None
        for fila_idx in range(fila_modulos + 1, min(fila_modulos + 15, df_raw.shape[0])):
            valor = df_raw.iloc[fila_idx, col_nombres]
            if pd.notna(valor) and ',' in str(valor) and len(str(valor)) > 10:
                fila_inicio = fila_idx
                break

        if not fila_inicio:
            return datos

        for fila_idx in range(fila_inicio, df_raw.shape[0]):
            nombre_valor = df_raw.iloc[fila_idx, col_nombres]

            if not pd.notna(nombre_valor):
                continue

            nombre = str(nombre_valor).strip().upper()
            if len(nombre) < 5 or ',' not in nombre:
                continue

            if nombre not in datos["alumnos"]:
                datos["alumnos"][nombre] = {}

            for info in modulos_info:
                if "nota_col" not in info or "calif_col" not in info:
                    continue

                modulo = info["modulo"]
                nota_val = df_raw.iloc[fila_idx, info["nota_col"]]
                calif_val = df_raw.iloc[fila_idx, info["calif_col"]]

                nota = str(nota_val).strip() if pd.notna(nota_val) and str(nota_val) != 'nan' else ""
                calif = str(calif_val).strip() if pd.notna(calif_val) and str(calif_val) != 'nan' else ""

                if nota:
                    if nota.lower() == 'baja':
                        nota = ""
                    else:
                        try:
                            n = float(nota)
                            nota = str(int(n))
                        except:
                            pass

                if calif:
                    if calif.lower() == 'baja':
                        calif = ""
                    else:
                        try:
                            c = float(calif)
                            calif = str(int(c))
                        except:
                            pass

                if nota and calif:
                    if nota.lower() == calif.lower():
                        datos["alumnos"][nombre][modulo] = nota
                    else:
                        datos["alumnos"][nombre][modulo] = f"{nota} {calif}"
                elif nota or calif:
                    valor = nota if nota else calif
                    datos["alumnos"][nombre][modulo] = valor

        st.success(f"Procesados {len(datos['alumnos'])} alumnos")
        return datos

    except Exception as e:
        st.error(f"Error: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return datos


def llenar_excel_resumen(excel_file, datos_excel, datos_documentos, datos_ctrl=None):
    """Combina datos del Excel, documentos escaneados y CTRL, y rellena la pestaña RESUMEN"""
    import openpyxl
    import re
    from datetime import datetime, date
    
    try:
        excel_file.seek(0)
        wb = openpyxl.load_workbook(excel_file)

        if "RESUMEN" not in wb.sheetnames:
            st.error("No se encontró la pestaña 'RESUMEN'")
            return None

        ws = wb["RESUMEN"]

        encabezados = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor:
                encabezados[str(valor).strip().lower()] = col

        alumnos_excel = datos_excel.get("alumnos", {})

        if not alumnos_excel:
            st.warning("No se encontraron alumnos en las pestañas del Excel")
            return None

        alumnos_lista = list(alumnos_excel.items())
        celdas_escritas = 0

        for i, (nombre, datos_alumno) in enumerate(alumnos_lista):
            fila = 2 + i

            if "id" in encabezados:
                ws.cell(row=fila, column=encabezados["id"], value=i + 1)
                celdas_escritas += 1

            if "nombre completo" in encabezados:
                ws.cell(row=fila, column=encabezados["nombre completo"], value=nombre)
                celdas_escritas += 1

            dni = ""
            if datos_ctrl:
                nombre_upper = nombre.upper()
                for nombre_ctrl, datos_ctrl_alumno in datos_ctrl.items():
                    if nombre_upper in nombre_ctrl or nombre_ctrl in nombre_upper:
                        dni_ctrl = datos_ctrl_alumno.get("dni", "")
                        if dni_ctrl:
                            dni = dni_ctrl
                            break

            if not dni:
                dni = datos_alumno.get("dni", "")

            if "dni" in encabezados and dni:
                ws.cell(row=fila, column=encabezados["dni"], value=dni)
                celdas_escritas += 1

            if datos_ctrl:
                nombre_upper = nombre.upper()
                datos_alumno_ctrl = None

                for nombre_ctrl, datos_ctrl_alumno in datos_ctrl.items():
                    if nombre_upper in nombre_ctrl or nombre_ctrl in nombre_upper:
                        datos_alumno_ctrl = datos_ctrl_alumno
                        break

                if datos_alumno_ctrl:
                    corporacion = datos_alumno_ctrl.get("corporacion_a_clase", "")
                    for enc_key, col in encabezados.items():
                        if "corporacion" in enc_key.lower() or "corporación" in enc_key.lower():
                            if corporacion:
                                ws.cell(row=fila, column=col, value=corporacion)
                                celdas_escritas += 1
                            break

                    baja = datos_alumno_ctrl.get("baja", "")
                    for enc_key, col in encabezados.items():
                        if enc_key.lower() == "baja":
                            if baja:
                                ws.cell(row=fila, column=col, value=baja)
                                celdas_escritas += 1
                            break

                    motivo = datos_alumno_ctrl.get("motivo_sin_parentesis", "")
                    for enc_key, col in encabezados.items():
                        if "motivo" in enc_key.lower() and "baja" not in enc_key.lower():
                            if motivo:
                                ws.cell(row=fila, column=col, value=motivo)
                                celdas_escritas += 1
                            break

                    baja_fecha_raw = datos_alumno_ctrl.get("baja", "")
                    motivo_sin_parentesis = datos_alumno_ctrl.get("motivo_sin_parentesis", "")

                    baja_fecha = ""
                    if baja_fecha_raw:
                        if isinstance(baja_fecha_raw, (datetime, date)):
                            baja_fecha = baja_fecha_raw.strftime('%d/%m/%Y')
                        else:
                            baja_str = str(baja_fecha_raw)
                            if '00:00:00' in baja_str:
                                try:
                                    fecha_obj = pd.to_datetime(baja_str)
                                    baja_fecha = fecha_obj.strftime('%d/%m/%Y')
                                except:
                                    baja_fecha = baja_str
                            else:
                                baja_fecha = baja_str

                    for enc_key, col in encabezados.items():
                        if ("baja" in enc_key.lower() and "motivo" in enc_key.lower()) or \
                           (enc_key.lower() == "baja - motivo"):
                            if baja_fecha and motivo_sin_parentesis:
                                baja_motivo_combinado = f"{baja_fecha} {motivo_sin_parentesis}"
                            elif baja_fecha:
                                baja_motivo_combinado = baja_fecha
                            elif motivo_sin_parentesis:
                                baja_motivo_combinado = motivo_sin_parentesis
                            else:
                                baja_motivo_combinado = ""

                            if baja_motivo_combinado:
                                ws.cell(row=fila, column=col, value=baja_motivo_combinado)
                                celdas_escritas += 1
                            break

                    baja_ocupacion = datos_alumno_ctrl.get("baja_ocupacion", "")
                    for enc_key, col in encabezados.items():
                        if ('baja' in enc_key.lower() and 'ocupacion' in enc_key.lower()) or \
                           ('baja' in enc_key.lower() and 'ocupación' in enc_key.lower()) or \
                           ('%' in enc_key and ('ocupacion' in enc_key.lower() or 'ocupación' in enc_key.lower())):
                            if baja_ocupacion:
                                ws.cell(row=fila, column=col, value=baja_ocupacion)
                                celdas_escritas += 1
                            break

                    fecha_incorporacion = datos_alumno_ctrl.get("fecha_incorporacion", "")

                    for enc_key, col in encabezados.items():
                        if ('incorporacion' in enc_key.lower() or 'incorporación' in enc_key.lower()) and \
                           ('clase' in enc_key.lower() or 'fecha' in enc_key.lower() or 'sintrafor' in enc_key.lower()):
                            if fecha_incorporacion:
                                celda = ws.cell(row=fila, column=col)

                                if isinstance(fecha_incorporacion, date):
                                    fecha_string = fecha_incorporacion.strftime('%d/%m/%Y')
                                    celda.value = fecha_string
                                    celda.number_format = '@'
                                elif isinstance(fecha_incorporacion, datetime):
                                    fecha_string = fecha_incorporacion.strftime('%d/%m/%Y')
                                    celda.value = fecha_string
                                    celda.number_format = '@'
                                else:
                                    celda.value = str(fecha_incorporacion)
                                    celda.number_format = '@'

                                celdas_escritas += 1
                            break

            porcentaje = datos_alumno.get("porcentaje_asistencia", "")

            for enc_key, col in encabezados.items():
                if "asistencia" in enc_key.lower() and "%" in enc_key:
                    if porcentaje:
                        ws.cell(row=fila, column=col, value=porcentaje)
                        celdas_escritas += 1
                    break

            fcoo03 = datos_alumno.get("fcoo03", "")

            fcoo03_col = None
            for enc_key, col in encabezados.items():
                if 'fcoo' in enc_key.lower() and '03' in enc_key.lower():
                    fcoo03_col = col
                    break

            if fcoo03_col and fcoo03:
                ws.cell(row=fila, column=fcoo03_col, value=fcoo03)
                celdas_escritas += 1

            modulos_mf = datos_alumno.get("modulos_mf", {})

            if modulos_mf:
                for modulo, calificacion in modulos_mf.items():
                    modulo_limpio = str(modulo).strip().upper()

                    for enc_key, col in encabezados.items():
                        enc_key_upper = enc_key.strip().upper()

                        if modulo_limpio == enc_key_upper or \
                           modulo_limpio.replace('_', '') == enc_key_upper.replace('_', '') or \
                           modulo_limpio in enc_key_upper or \
                           enc_key_upper in modulo_limpio:
                            ws.cell(row=fila, column=col, value=calificacion)
                            celdas_escritas += 1
                            break

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    except Exception as e:
        st.error(f"Error: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return None


def render_formacion_fin():
    """Renderiza la sección de Formación Empresa Fin"""
    st.markdown("### Fase 1: Cargar Archivos Excel")
    st.markdown('<div class="custom-card"><p>Sube los archivos Excel necesarios para completar el proceso de documentación.</p></div>', unsafe_allow_html=True)

    col_excel1, col_excel2 = st.columns(2)

    with col_excel1:
        st.markdown("**Excel Principal** (Resumen, Calificaciones, Asistencia)")
        excel_justificacion = st.file_uploader(
            "Selecciona el archivo Excel",
            type=["xlsx", "xls"],
            key="excel_justificacion",
            help="Arrastra aquí tu archivo Excel o haz clic para seleccionar"
        )

        if excel_justificacion:
            st.success("Excel principal cargado")
            try:
                xls = pd.ExcelFile(excel_justificacion)
                st.write(f"Pestañas: {', '.join(xls.sheet_names)}")
            except Exception as e:
                st.warning(f"Error: {str(e)}")

    with col_excel2:
        st.markdown("**Excel CTRL de Alumnos** (Pestaña CTRL)")
        excel_ctrl = st.file_uploader(
            "Selecciona el archivo Excel CTRL",
            type=["xlsx", "xls"],
            key="excel_ctrl",
            help="Este Excel debe contener la pestaña CTRL con información de corporación, baja y motivo"
        )

        if excel_ctrl:
            st.success("Excel CTRL cargado")
            try:
                xls_ctrl = pd.ExcelFile(excel_ctrl)
                st.write(f"Pestañas: {', '.join(xls_ctrl.sheet_names)}")
                if "CTRL" in xls_ctrl.sheet_names:
                    st.info("Pestaña CTRL encontrada")
                else:
                    st.warning("No se encontró la pestaña CTRL")
            except Exception as e:
                st.warning(f"Error: {str(e)}")

    st.markdown("---")
    st.markdown("### Fase 2: Documentos Escaneados (Requeridos)")
    st.markdown('<div class="custom-card"><p>Carga los documentos necesarios. Puedes cargar múltiples hojas de firmas.</p></div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**Plan de Evaluación**")
        plan_evaluacion = st.file_uploader(
            "Selecciona el Plan de Evaluación",
            key="plan_evaluacion",
            help="Arrastra aquí tu archivo o haz clic para seleccionar"
        )

        st.markdown("**Cronograma**")
        cronograma = st.file_uploader(
            "Selecciona el Cronograma",
            key="cronograma",
            help="Arrastra aquí tu archivo o haz clic para seleccionar"
        )

    with col2:
        st.markdown("**Certificados de Asistencia (Hojas de Firmas)**")
        certificados = st.file_uploader(
            "Selecciona una o más Hojas de Firmas",
            key="certificados",
            accept_multiple_files=True,
            help="Puedes seleccionar múltiples archivos. Arrastra aquí o haz clic para seleccionar"
        )
        
        if certificados:
            st.success(f"{len(certificados)} archivo(s) de hojas de firmas cargado(s)")
            for cert in certificados:
                st.write(f"- {cert.name}")

        st.markdown("**Evaluación de Profesores**")
        evacuacion = st.file_uploader(
            "Selecciona el archivo de Evaluación",
            key="evacuacion",
            help="Arrastra aquí tu archivo o haz clic para seleccionar"
        )

    st.markdown("---")
    st.markdown("### Fase 3: Procesamiento Automático")

    archivos_totales = sum(1 for doc in [plan_evaluacion, cronograma, evacuacion] if doc is not None) + (1 if certificados else 0)

    if archivos_totales > 0:
        st.markdown(f'<div class="custom-card"><p><strong>{archivos_totales} de 4</strong> tipos de documentos cargados correctamente</p></div>', unsafe_allow_html=True)

    excel_ctrl_cargado = excel_ctrl is not None

    if excel_justificacion and plan_evaluacion and cronograma and certificados and evacuacion:

        if not excel_ctrl_cargado:
            st.warning("AVISO: Excel CTRL no cargado. Los campos de Corporación, Baja y Motivo quedarán vacíos.")

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Procesar y Completar Resumen", type="primary", use_container_width=True):

            progress_bar = st.progress(0)
            status_text = st.empty()

            datos_ctrl = None
            if excel_ctrl_cargado:
                status_text.text("Procesando Excel CTRL...")
                progress_bar.progress(10)
                datos_ctrl = leer_datos_ctrl(excel_ctrl)

            datos_documentos = {}
            datos_evaluacion = None

            status_text.text("Procesando Plan de Evaluación...")
            progress_bar.progress(20)
            if plan_evaluacion:
                texto_plan = procesar_documento(plan_evaluacion)

            status_text.text("Procesando Cronograma...")
            progress_bar.progress(30)
            if cronograma:
                texto_cronograma = procesar_documento(cronograma)

            status_text.text("Procesando Hojas de Firmas...")
            progress_bar.progress(40)
            if certificados:
                todos_alumnos = []
                for certificado in certificados:
                    texto_certificado = procesar_documento(certificado)
                    if texto_certificado:
                        datos_cert = extraer_datos_certificado_asistencia(texto_certificado)
                        if datos_cert.get("alumnos"):
                            todos_alumnos.extend(datos_cert["alumnos"])
                
                if todos_alumnos:
                    datos_documentos["certificado_asistencia"] = {
                        "alumnos": todos_alumnos,
                        "fecha_inicio": "",
                        "curso": ""
                    }

            status_text.text("Procesando Evaluación de Profesores...")
            progress_bar.progress(50)
            if evacuacion:
                if evacuacion.type in ["application/vnd.ms-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
                    evacuacion.seek(0)
                    datos_evaluacion = extraer_evaluacion_excel(evacuacion, verbose=False)
                else:
                    texto_evacuacion = procesar_documento(evacuacion)
                    if texto_evacuacion:
                        datos_evaluacion = extraer_evaluacion_profesores(texto_evacuacion)

            status_text.text("Leyendo datos del Excel principal...")
            progress_bar.progress(70)
            datos_excel = leer_datos_excel(excel_justificacion, datos_evaluacion)

            status_text.text("Completando resumen...")
            progress_bar.progress(90)

            try:
                excel_justificacion.seek(0)
                excel_bytes = excel_justificacion.read()
                excel_buffer = io.BytesIO(excel_bytes)

                excel_completado = llenar_excel_resumen(excel_buffer, datos_excel, datos_documentos, datos_ctrl)
            except Exception as e:
                st.error(f"ERROR: {str(e)}")
                excel_completado = None

            progress_bar.progress(100)
            status_text.empty()
            progress_bar.empty()

            if excel_completado:
                st.markdown("""
                <div class="result-container">
                    <h2 class="result-title">Proceso Completado</h2>
                    <p style="color: #065f46; font-size: 1rem;">El resumen ha sido generado exitosamente con todos los datos integrados.</p>
                </div>
                """, unsafe_allow_html=True)

                col1, col2, col3, col4 = st.columns(4)

                with col1:
                    st.metric("Alumnos", len(datos_excel.get('alumnos', {})))

                with col2:
                    evaluaciones = sum(len(mods) for mods in datos_evaluacion.get("alumnos", {}).values()) if datos_evaluacion else 0
                    st.metric("Calificaciones", evaluaciones)

                with col3:
                    ctrl_status = "SÍ" if datos_ctrl else "NO"
                    st.metric("Excel CTRL", ctrl_status)

                with col4:
                    docs_procesados = sum([
                        1 if plan_evaluacion else 0,
                        1 if cronograma else 0,
                        1 if certificados else 0,
                        1 if evacuacion else 0
                    ])
                    st.metric("Documentos", f"{docs_procesados}/4")

                st.markdown("<br>", unsafe_allow_html=True)

                col_btn1, col_btn2 = st.columns([2, 1])
                with col_btn1:
                    st.download_button(
                        label="Descargar Excel Completado",
                        data=excel_completado,
                        file_name=f"resumen_completado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )

                st.markdown("""
                <div class="custom-card">
                    <h3>Datos Integrados</h3>
                    <p><strong>Información incorporada en el resumen:</strong></p>
                    <ul style="color: var(--text-secondary); line-height: 1.8;">
                        <li>Datos de alumnos del Excel principal y CTRL</li>
                        <li>Fechas de incorporación y motivos de baja</li>
                        <li>DNI y corporación de cada alumno</li>
                        <li>Porcentajes de asistencia (MF+FCOO+FE)</li>
                        <li>Calificaciones de módulos (MF) de la evaluación</li>
                        <li>Calificaciones de FCOO03</li>
                    </ul>
                    <p style="margin-top: 1rem; font-weight: 500;">Descarga el archivo y revisa que toda la información sea correcta.</p>
                </div>
                """, unsafe_allow_html=True)

            else:
                st.error("ERROR: Hubo un error al completar el Excel. Por favor, revisa los archivos cargados.")

    elif excel_justificacion and archivos_totales > 0:
        st.warning("AVISO: Por favor, carga TODOS los documentos necesarios para continuar")
        tipos_faltantes = []
        if not plan_evaluacion:
            tipos_faltantes.append("Plan de Evaluación")
        if not cronograma:
            tipos_faltantes.append("Cronograma")
        if not certificados:
            tipos_faltantes.append("Hojas de Firmas")
        if not evacuacion:
            tipos_faltantes.append("Evaluación de Profesores")
        
        st.info(f"Documentos faltantes: {', '.join(tipos_faltantes)}")