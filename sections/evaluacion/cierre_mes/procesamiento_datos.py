"""
VERSION FINAL - Camelot con observaciones corregidas
"""

import re
from collections import defaultdict
from datetime import datetime
import os

os.environ['TESSDATA_PREFIX'] = r'C:\Program Files\Tesseract-OCR\tessdata'
POPPLER_PATH = r'C:\Users\Arancha\Downloads\poppler-24.08.0\Library\bin'


def extraer_becas_ayudas_tabla(pdf_path, alumnos_excel=None):
    """
    Extrae becas usando camelot - Lee columnas de la tabla
    """
    import camelot
    import pandas as pd
    
    if not alumnos_excel:
        print("ERROR: Se necesita lista de alumnos con DNI")
        return {}
    
    ayudas_dict = {}
    
    print("=" * 80)
    print("EXTRAYENDO BECAS CON CAMELOT")
    print("=" * 80)
    
    try:
        # Extraer tabla
        print("\nExtrayendo tabla del PDF...")
        
        tables = camelot.read_pdf(
            pdf_path,
            pages='1',
            flavor='stream',
            table_areas=['50,500,800,100']
        )
        
        if len(tables) == 0:
            print("ERROR: No se encontraron tablas")
            return {}
        
        df = tables[0].df
        
        print(f"Tabla: {df.shape[0]} filas x {df.shape[1]} columnas\n")
        
        print("="*80)
        print("BUSCANDO ALUMNOS")
        print("="*80)
        
        encontrados = 0
        
        for alumno in alumnos_excel:
            nombre = alumno['nombre_completo']
            dni = alumno['dni']
            
            # Buscar DNI en la tabla
            fila_encontrada = None
            
            for idx in range(len(df)):
                fila = df.iloc[idx]
                fila_texto = ' '.join([str(x) for x in fila if pd.notna(x)])
                
                if dni in fila_texto:
                    fila_encontrada = fila
                    break
            
            ayudas = []
            
            if fila_encontrada is not None:
                print(f"\n{nombre} ({dni}):")
                
                # Estructura de columnas:
                # 0: Nº
                # 1: Apellidos y Nombre
                # 2: NIF
                # 3: €/día
                # 4: Ayuda transporte (X)
                # 5: Manutención
                # 6: Alojamiento/Beca (X para Discapacidad)
                # 7: Ayuda conciliación (X)
                # 8: Otras (texto "DISCAPACIDAD")
                
                # DISCAPACIDAD
                col_8 = str(fila_encontrada.iloc[8] if len(fila_encontrada) > 8 else '').upper()
                col_6 = str(fila_encontrada.iloc[6] if len(fila_encontrada) > 6 else '').strip().upper()
                
                if 'DISCAPAC' in col_8 or col_6 == 'X':
                    ayudas.append('Discapacidad')
                    print(f"  -> Discapacidad")
                else:
                    # TRANSPORTE (columna 4)
                    col_4 = str(fila_encontrada.iloc[4] if len(fila_encontrada) > 4 else '').strip().upper()
                    
                    # CONCILIACION (columna 7)
                    col_7 = str(fila_encontrada.iloc[7] if len(fila_encontrada) > 7 else '').strip().upper()
                    
                    if col_4 == 'X':
                        ayudas.append('Transporte')
                    
                    if col_7 == 'X':
                        ayudas.append('Conciliación')
                    
                    if ayudas:
                        print(f"  -> {' + '.join(ayudas)}")
                    else:
                        print(f"  -> Sin ayudas")
                
                encontrados += 1
            else:
                print(f"\n{nombre} ({dni}): NO ENCONTRADO")
            
            ayudas_dict[nombre] = ayudas
        
        print(f"\n{'=' * 80}")
        print(f"Encontrados: {encontrados}/{len(alumnos_excel)}")
        print(f"Con ayudas: {len([a for a in ayudas_dict.values() if a])}")
        print(f"{'=' * 80}\n")
        
    except ImportError:
        print("\nERROR: camelot-py no instalado")
        print("pip install camelot-py[cv] --break-system-packages")
        return {}
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
    
    return ayudas_dict


extraer_becas_por_filas_imagen = extraer_becas_ayudas_tabla


def extraer_alumnos_excel(excel_path):
    import openpyxl
    alumnos = []
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        nombre_col = None
        dni_col = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value:
                    val = str(cell.value).lower()
                    if 'nombre' in val or 'apellido' in val:
                        nombre_col = cell.column
                    if 'dni' in val or 'nif' in val:
                        dni_col = cell.column
        if not nombre_col or not dni_col:
            return []
        for row in ws.iter_rows(min_row=2):
            nombre = row[nombre_col - 1].value
            dni = row[dni_col - 1].value
            if nombre and dni:
                nombre_str = str(nombre).strip()
                dni_str = str(dni).strip()
                if 'apellido' in nombre_str.lower() or nombre_str.lower() == 'nombre':
                    continue
                if not re.match(r'^[Z0-9]\d{7}[A-Z]$', dni_str):
                    continue
                alumnos.append({'nombre_completo': nombre_str, 'dni': dni_str})
        print(f"\n{len(alumnos)} alumnos del Excel\n")
    except Exception as e:
        print(f"Error: {e}")
    return alumnos


def calcular_dias_lectivos_y_faltas_corregido(firmas_pdfs, alumnos_excel, dias_asistidos_manual=None):
    dias_lectivos = 23
    print(f"\n{'=' * 80}")
    print(f"DIAS LECTIVOS: {dias_lectivos} (21 empresa + 2 aula)")
    print(f"{'=' * 80}\n")
    print("CALCULANDO FALTAS:\n")
    dias_asistidos_dict = {}
    faltas_dict = {}
    for alumno in alumnos_excel:
        nombre = alumno['nombre_completo']
        if dias_asistidos_manual and nombre in dias_asistidos_manual:
            dias_asistidos = dias_asistidos_manual[nombre]
        else:
            dias_asistidos = 23
        faltas = dias_lectivos - dias_asistidos
        dias_asistidos_dict[nombre] = dias_asistidos
        faltas_dict[nombre] = faltas
        print(f"{nombre}")
        print(f"  Dias asistidos: {dias_asistidos}")
        print(f"  Faltas: {faltas}\n")
    print(f"{'=' * 80}\n")
    return dias_lectivos, faltas_dict, dias_asistidos_dict


def extraer_justificantes_mejorado(pdf_path):
    import pytesseract
    from pdf2image import convert_from_path
    justificantes_dict = defaultdict(int)
    try:
        print("=" * 80)
        print("EXTRAYENDO JUSTIFICANTES")
        print("=" * 80)
        images = convert_from_path(pdf_path, dpi=300, poppler_path=POPPLER_PATH)
        for image in images:
            try:
                osd = pytesseract.image_to_osd(image)
                angle = int(re.search(r'Rotate: (\d+)', osd).group(1))
                if angle != 0:
                    image = image.rotate(-angle, expand=True)
            except:
                pass
            texto = pytesseract.image_to_string(image, lang='spa', config='--psm 6')
            matches = re.finditer(r'([A-ZÁÉÍÓÚÑ]+(?:\s+[A-ZÁÉÍÓÚÑ]+)*)[,\s]+([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)*)', texto)
            nombres_en_pagina = set()
            for match in matches:
                apellidos = match.group(1).strip().upper()
                nombre_pila = match.group(2).strip().upper()
                nombre_pila_formateado = ' '.join([p.capitalize() for p in nombre_pila.split()])
                nombre_completo = f"{apellidos}, {nombre_pila_formateado}"
                if len(nombre_completo) > 10 and nombre_completo not in nombres_en_pagina:
                    nombres_en_pagina.add(nombre_completo)
                    justificantes_dict[nombre_completo] += 1
                    print(f"  {nombre_completo}: +1 justificante")
        print(f"\nTotal: {len(justificantes_dict)} alumnos con justificantes\n")
    except Exception as e:
        print(f"Error: {e}")
    return dict(justificantes_dict)


def construir_observaciones_completas(nombre_alumno, ayudas_dict, dias_asistidos_dict, justificantes_dict):
    """Construye observaciones con formato correcto"""
    lineas = []
    
    # Línea 1: Ayudas (Transporte + Conciliación: 19)
    ayudas = ayudas_dict.get(nombre_alumno, [])
    dias = dias_asistidos_dict.get(nombre_alumno, 0)
    
    if ayudas and dias > 0:
        # Formatear ayudas con capitalización correcta
        ayudas_formateadas = []
        for ayuda in ayudas:
            if ayuda.upper() == 'DISCAPACIDAD':
                ayudas_formateadas.append('Discapacidad')
            elif ayuda.upper() == 'TRANSPORTE':
                ayudas_formateadas.append('Transporte')
            elif ayuda.upper() == 'CONCILIACIÓN' or ayuda.upper() == 'CONCILIACION':
                ayudas_formateadas.append('Conciliación')
            else:
                ayudas_formateadas.append(ayuda.capitalize())
        
        ayudas_texto = ' + '.join(ayudas_formateadas)
        lineas.append(f"{ayudas_texto}: {dias}")
    
    # Línea 2: Justificantes (1 falta justificada)
    justif = justificantes_dict.get(nombre_alumno, 0)
    if justif > 0:
        texto_justif = f"{justif} falta{'s' if justif > 1 else ''} justificada{'s' if justif > 1 else ''}"
        lineas.append(texto_justif)
    
    return '\n'.join(lineas) if lineas else ''


def obtener_mes_anterior():
    meses = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
             'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    mes_actual = datetime.now().month
    mes_anterior = (mes_actual - 2) % 12
    return meses[mes_anterior]


def extraer_datos_curso_pdf(pdf_path):
    import pytesseract
    from pdf2image import convert_from_path
    try:
        images = convert_from_path(pdf_path, dpi=300, poppler_path=POPPLER_PATH)
        for image in images:
            texto = pytesseract.image_to_string(image, lang='spa', config='--psm 6')
            match_curso = re.search(r'N[°º]\s*de\s*Curso[:\s]+(\d{4}/\d+)', texto, re.IGNORECASE)
            numero_curso = match_curso.group(1) if match_curso else ''
            match_esp = re.search(r'OPERACIONES AUXILIARES[^\n]+', texto, re.IGNORECASE)
            especialidad = match_esp.group(0) if match_esp else ''
            if numero_curso or especialidad:
                return numero_curso, especialidad
        return '', ''
    except:
        return '', ''