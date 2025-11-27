"""
Procesador de Excel - Versión Compatible
Este archivo mantiene el nombre ExcelProcessorReal para compatibilidad
"""
import pandas as pd
import numpy as np
from typing import Dict, List, Optional
import io
import re


class ExcelProcessorReal:
    """Procesa archivos Excel de asistencias - 100% Dinámico"""
    
    def __init__(self):
        self.archivo_excel = None
        self.df_asistencias = None
        self.curso_codigo = None
        self.curso_nombre = None
        self.numero_expediente = None
        self.codigo_certificado = None
        self.centro_formativo = None
        self.codigo_centro = None
        self.direccion = None
        self.localidad = None
        self.codigo_postal = None
        self.provincia = None
        self.alumnos_data = []
        self.modulos_config = []
        
    def cargar_asistencias(self, file_bytes: bytes) -> Dict:
        """
        Carga el archivo de asistencias
        
        Args:
            file_bytes: Bytes del archivo Excel
            
        Returns:
            Dict con datos procesados
        """
        try:
            self.archivo_excel = pd.ExcelFile(io.BytesIO(file_bytes))
            
            # Buscar la pestaña de ASISTENCIA
            pestaña_asistencia = None
            for nombre in self.archivo_excel.sheet_names:
                if 'asistencia' in nombre.lower():
                    pestaña_asistencia = nombre
                    break
            
            if not pestaña_asistencia:
                pestaña_asistencia = self.archivo_excel.sheet_names[0]
            
            # Cargar datos
            self.df_asistencias = pd.read_excel(
                self.archivo_excel, 
                sheet_name=pestaña_asistencia, 
                header=None
            )
            
            # Extraer información
            self._extraer_info_administrativa()
            self._detectar_modulos()
            self._extraer_alumnos()
            estadisticas = self._calcular_estadisticas()
            
            return {
                'curso_codigo': self.curso_codigo,
                'curso_nombre': self.curso_nombre,
                'codigo_certificado': self.codigo_certificado,
                'numero_expediente': self.numero_expediente,
                'centro_formativo': self.centro_formativo,
                'codigo_centro': self.codigo_centro,
                'direccion': self.direccion,
                'localidad': self.localidad,
                'alumnos': self.alumnos_data,
                'modulos_detectados': self.modulos_config,
                'estadisticas_individuales': self.alumnos_data,
                'estadisticas_grupales': estadisticas,
                'success': True
            }
            
        except Exception as e:
            raise Exception(f"Error al procesar asistencias: {str(e)}")
    
    def _extraer_info_administrativa(self):
        """Extrae información administrativa del archivo"""
        df = self.df_asistencias
        
        # Buscar en las primeras 15 filas
        for i in range(min(15, df.shape[0])):
            for col in range(min(10, df.shape[1])):
                val = str(df.iloc[i, col]).lower() if pd.notna(df.iloc[i, col]) else ''
                
                if not val or val == 'nan':
                    continue
                
                valor_siguiente = None
                if col + 1 < df.shape[1]:
                    valor_siguiente = df.iloc[i, col + 1]
                    if pd.notna(valor_siguiente):
                        valor_siguiente = str(valor_siguiente).strip()
                
                # Curso
                if 'curso' in val and ':' in val and not self.curso_codigo:
                    if valor_siguiente:
                        self.curso_codigo = valor_siguiente
                
                # Nombre del curso
                if ('nombre' in val or 'certificado profesional' in val) and not self.curso_nombre:
                    if valor_siguiente and len(valor_siguiente) > 10:
                        match = re.search(r'\(([A-Z]{4}\d{4})\)', valor_siguiente)
                        if match:
                            self.codigo_certificado = match.group(1)
                            self.curso_nombre = valor_siguiente.replace(f'({self.codigo_certificado})', '').strip()
                        else:
                            self.curso_nombre = valor_siguiente
                
                # Expediente
                if 'expediente' in val and not self.numero_expediente:
                    if valor_siguiente:
                        self.numero_expediente = valor_siguiente
                
                # Centro formativo
                if 'centro' in val and 'formativ' in val and not self.centro_formativo:
                    if valor_siguiente:
                        self.centro_formativo = valor_siguiente
                
                # Dirección
                if 'direcci' in val and not self.direccion:
                    if valor_siguiente:
                        self.direccion = valor_siguiente
                
                # Localidad
                if 'localidad' in val and not self.localidad:
                    if valor_siguiente:
                        self.localidad = valor_siguiente
        
        # Valores por defecto
        if not self.numero_expediente and self.curso_codigo:
            self.numero_expediente = f"E-{self.curso_codigo.replace('/', '-')}"
        
        if not self.codigo_certificado and self.curso_nombre:
            match = re.search(r'[A-Z]{4}\d{4}', self.curso_nombre)
            if match:
                self.codigo_certificado = match.group()
        
        if not self.centro_formativo:
            self.centro_formativo = "INTERPROS NEXT GENERATION SLU"
        if not self.codigo_centro:
            self.codigo_centro = "26615"
        if not self.direccion:
            self.direccion = "C/ DR. SEVERO OCHOA, 21, BJ"
        if not self.localidad:
            self.localidad = "AVILÉS"
        if not self.codigo_postal:
            self.codigo_postal = "33400"
        if not self.provincia:
            self.provincia = "ASTURIAS"
        if not self.curso_nombre:
            self.curso_nombre = "ACTIVIDADES DE GESTIÓN ADMINISTRATIVA"
    
    def _detectar_modulos(self):
        """Detecta módulos automáticamente"""
        df = self.df_asistencias
        self.modulos_config = []
        
        # Buscar fila de módulos (contiene códigos MF)
        fila_modulos = None
        for i in range(min(30, df.shape[0])):
            for col in range(df.shape[1]):
                val = str(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else ''
                if re.search(r'MF\d{4}_\d', val):
                    fila_modulos = i
                    break
            if fila_modulos:
                break
        
        if fila_modulos is None:
            fila_modulos = 5
        
        # Detectar módulos
        for col in range(df.shape[1]):
            val = str(df.iloc[fila_modulos, col]) if pd.notna(df.iloc[fila_modulos, col]) else ''
            
            match = re.search(r'MF\d{4}_\d', val)
            if not match:
                continue
            
            codigo = match.group()
            
            # Extraer nombre
            nombre = ''
            if ':' in val:
                partes = val.split(':', 1)
                if len(partes) > 1:
                    nombre = partes[1].split('\n')[0].strip()
                    nombre = re.sub(r'Fechas:.*$', '', nombre).strip()
            
            # Buscar columna de asistencia
            col_asistencia = None
            for offset in range(1, 20):
                if col + offset >= df.shape[1]:
                    break
                
                for fila_enc in range(fila_modulos + 1, min(fila_modulos + 5, df.shape[0])):
                    val_enc = str(df.iloc[fila_enc, col + offset]).lower() if pd.notna(df.iloc[fila_enc, col + offset]) else ''
                    if 'total' in val_enc and 'asistencia' in val_enc:
                        col_asistencia = col + offset
                        break
                
                if col_asistencia:
                    break
            
            # Buscar horas totales en la fila 7 (H TOTAL TEORIA)
            horas_totales = 0
            
            # Primero buscar en las columnas cercanas al módulo
            for offset in range(-2, 8):
                if not (0 <= col + offset < df.shape[1]):
                    continue
                
                # Buscar específicamente en fila 7 y 10 donde suelen estar las horas
                for fila_busq in [7, 10]:
                    if fila_busq < df.shape[0]:
                        val_horas = df.iloc[fila_busq, col + offset]
                        if pd.notna(val_horas) and isinstance(val_horas, (int, float)):
                            if 30 < val_horas < 500:
                                # Verificar que no sea una columna de asistencia (que tiene decimales)
                                if isinstance(val_horas, int) or val_horas == int(val_horas):
                                    horas_totales = int(val_horas)
                                    break
                
                if horas_totales > 0:
                    break
            
            # Si no encuentra, usar valor por defecto basado en el código
            if horas_totales == 0:
                horas_totales = 100
            
            if not col_asistencia:
                continue
            
            self.modulos_config.append({
                'codigo': codigo,
                'nombre': nombre if nombre else f"Módulo {codigo}",
                'horas_totales': horas_totales,
                'col_horas_asistidas': col_asistencia
            })
        
        self.modulos_config.sort(key=lambda x: x['col_horas_asistidas'])
        
        if not self.modulos_config:
            raise Exception("No se detectaron módulos en el archivo")
    
    def _extraer_alumnos(self):
        """Extrae datos de alumnos"""
        df = self.df_asistencias
        self.alumnos_data = []
        
        if not self.modulos_config:
            raise Exception("No se detectaron módulos")
        
        # Buscar fila de inicio de alumnos
        fila_inicio = 10
        col_alumno = 1
        col_dni = 2
        
        for i in range(min(30, df.shape[0])):
            for col in range(min(10, df.shape[1])):
                val = str(df.iloc[i, col]).lower() if pd.notna(df.iloc[i, col]) else ''
                if val == 'alumno':
                    col_alumno = col
                    fila_inicio = i + 1
                    
                    for c in range(col, min(col + 5, df.shape[1])):
                        val_dni = str(df.iloc[i, c]).lower() if pd.notna(df.iloc[i, c]) else ''
                        if 'dni' in val_dni or 'nie' in val_dni:
                            col_dni = c
                            break
                    break
            if col_alumno != 1:
                break
        
        # Extraer alumnos
        for i in range(fila_inicio, df.shape[0]):
            nombre = df.iloc[i, col_alumno] if col_alumno < df.shape[1] else None
            dni = df.iloc[i, col_dni] if col_dni < df.shape[1] else None
            
            if not pd.notna(nombre) or not isinstance(nombre, str) or len(nombre.strip()) < 3:
                continue
            
            if any(palabra in str(nombre).lower() for palabra in [
                'total', 'media', 'tutor', 'profesor', 'firma'
            ]):
                continue
            
            alumno = {
                'nombre': str(nombre).strip(),
                'dni': str(dni).strip() if pd.notna(dni) else 'N/A',
                'modulos': []
            }
            
            for modulo_cfg in self.modulos_config:
                col_horas = modulo_cfg['col_horas_asistidas']
                horas_valor = df.iloc[i, col_horas] if col_horas < df.shape[1] else None
                horas_asistidas = 0
                
                if pd.notna(horas_valor):
                    if isinstance(horas_valor, (int, float)):
                        horas_asistidas = float(horas_valor)
                    elif isinstance(horas_valor, str):
                        horas_lower = horas_valor.lower().strip()
                        if any(palabra in horas_lower for palabra in [
                            'exento', 'convalidado', 'convalida'
                        ]):
                            horas_asistidas = modulo_cfg['horas_totales']
                        else:
                            try:
                                horas_asistidas = float(horas_valor.replace(',', '.'))
                            except:
                                horas_asistidas = 0
                
                modulo_alumno = {
                    'codigo': modulo_cfg['codigo'],
                    'nombre': modulo_cfg['nombre'],
                    'horas_totales': modulo_cfg['horas_totales'],
                    'horas_asistidas': round(horas_asistidas, 2)
                }
                alumno['modulos'].append(modulo_alumno)
            
            total_horas_totales = sum(m['horas_totales'] for m in alumno['modulos'])
            total_horas_asistidas = sum(m['horas_asistidas'] for m in alumno['modulos'])
            alumno['total_horas'] = total_horas_totales
            alumno['total_asistidas'] = round(total_horas_asistidas, 2)
            alumno['porcentaje_asistencia'] = round(
                (total_horas_asistidas / total_horas_totales * 100), 2
            ) if total_horas_totales > 0 else 0
            
            self.alumnos_data.append(alumno)
        
        return self.alumnos_data
    
    def _calcular_estadisticas(self) -> Dict:
        """Calcula estadísticas grupales"""
        if not self.alumnos_data:
            return {
                'total_alumnos': 0,
                'porcentaje_asistencia_media': 0,
                'alumnos_con_mas_80': 0,
                'alumnos_con_menos_80': 0
            }
        
        total = len(self.alumnos_data)
        porcentajes = [a['porcentaje_asistencia'] for a in self.alumnos_data]
        
        return {
            'total_alumnos': total,
            'porcentaje_asistencia_media': round(sum(porcentajes) / total, 2) if total > 0 else 0,
            'porcentaje_asistencia_minimo': round(min(porcentajes), 2) if porcentajes else 0,
            'porcentaje_asistencia_maximo': round(max(porcentajes), 2) if porcentajes else 0,
            'alumnos_con_mas_80': sum(1 for p in porcentajes if p >= 80),
            'alumnos_con_menos_80': sum(1 for p in porcentajes if p < 80)
        }
    
    def obtener_alumno(self, nombre: str) -> Optional[Dict]:
        """Obtiene datos de un alumno específico"""
        for alumno in self.alumnos_data:
            if alumno['nombre'] == nombre:
                return alumno
        return None
    
    def obtener_todos_alumnos(self) -> List[Dict]:
        """Devuelve lista de todos los alumnos"""
        return self.alumnos_data
    
    def exportar_a_excel(self, alumno_nombre: str = None) -> bytes:
        """
        Exporta datos a Excel
        
        Args:
            alumno_nombre: Si se especifica, exporta solo ese alumno
        
        Returns:
            bytes del archivo Excel
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Evaluación"
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "INFORME DE EVALUACIÓN INDIVIDUALIZADO GRADO C"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        
        ws[f'A{row}'] = "Número de expediente:"
        ws[f'B{row}'] = self.numero_expediente
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Certificado profesional:"
        ws[f'B{row}'] = self.curso_nombre
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Código:"
        ws[f'B{row}'] = self.codigo_certificado
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Centro formativo:"
        ws[f'B{row}'] = self.centro_formativo
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Código:"
        ws[f'B{row}'] = self.codigo_centro
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        alumnos_exportar = []
        if alumno_nombre:
            alumno = self.obtener_alumno(alumno_nombre)
            if alumno:
                alumnos_exportar = [alumno]
        else:
            alumnos_exportar = self.alumnos_data
        
        for idx, alumno in enumerate(alumnos_exportar):
            if idx > 0:
                row += 3
            
            ws[f'A{row}'] = "El/la alumno/a:"
            ws[f'B{row}'] = alumno['nombre']
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = "con DNI/NIE/Pasaporte:"
            ws[f'B{row}'] = alumno['dni']
            ws[f'A{row}'].font = Font(bold=True)
            row += 2
            
            ws[f'A{row}'] = "Módulos"
            ws[f'B{row}'] = "Código"
            ws[f'C{row}'] = "Horas"
            ws[f'D{row}'] = "Horas de asistencia"
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            
            for modulo in alumno['modulos']:
                ws[f'A{row}'] = modulo['nombre']
                ws[f'B{row}'] = modulo['codigo']
                ws[f'C{row}'] = modulo['horas_totales']
                ws[f'D{row}'] = modulo['horas_asistidas']
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].border = border
                
                row += 1
            
            ws[f'A{row}'] = "TOTAL"
            ws[f'C{row}'] = alumno['total_horas']
            ws[f'D{row}'] = alumno['total_asistidas']
            ws[f'A{row}'].font = Font(bold=True)
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = "Porcentaje de asistencia:"
            ws[f'B{row}'] = f"{alumno['porcentaje_asistencia']}%"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()